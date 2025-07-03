import os
import openpyxl
from datetime import datetime
from telegram import Update, InlineKeyboardMarkup, InlineKeyboardButton, InputFile
from telegram.ext import (
    ApplicationBuilder, CommandHandler, MessageHandler, filters,
    ContextTypes, ConversationHandler, CallbackQueryHandler
)
import matplotlib.pyplot as plt
from io import BytesIO

# Pega o token e chat_id das variáveis de ambiente
TOKEN = os.getenv("TOKEN")
CHAT_ID = int(os.getenv("CHAT_ID"))

# Estados da conversa
(
    VALOR, DATA_RESULTADO, NUM_CASAS, CASAS, ODDS, VALOR_INVESTIDO,
    EDIT_SELECTION, EDIT_FIELD, EDIT_VALUE
) = range(9)

user_data = {}

def criar_planilha():
    file_name = "surebets.xlsx"
    if not os.path.exists(file_name):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "SureBets"
        ws.append([
            "Data Aposta", "Data Resultado",
            "Casa 1", "Odd 1",
            "Casa 2", "Odd 2",
            "Casa 3", "Odd 3",
            "Lucro Casa 1", "Lucro Casa 2", "Lucro Casa 3",
            "Valor Investido", "Lucro Total"
        ])
        wb.save(file_name)

def calcular_lucro(odds, valor_investido):
    lucros = []
    for odd in odds:
        if odd and odd > 0:
            lucro = valor_investido * odd - valor_investido
            lucros.append(round(lucro, 2))
        else:
            lucros.append(0)
    return lucros

def formatar_mensagem(aposta):
    # aposta é lista com colunas da planilha numa linha
    casas = [aposta[2], aposta[4], aposta[6]]
    odds = [aposta[3], aposta[5], aposta[7]]
    lucros = [aposta[8], aposta[9], aposta[10]]
    valor = aposta[11]
    lucro_total = aposta[12]
    data_aposta = aposta[0]
    data_resultado = aposta[1]

    texto = f"*Data aposta:* {data_aposta}\n*Data resultado:* {data_resultado}\n"
    texto += f"*Valor investido:* R$ {valor:.2f}\n"
    texto += f"*Lucro total esperado:* R$ {lucro_total:.2f}\n\n"

    for i in range(3):
        if casas[i]:
            texto += f"🏠 *Casa {i+1}:* {casas[i]}\n"
            texto += f"  - Odd: {odds[i]}\n"
            texto += f"  - Lucro esperado: R$ {lucros[i]:.2f}\n\n"
    return texto, lucro_total < 0

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "Olá! Use /nova para registrar uma nova aposta.\n"
        "Use /historico para ver as apostas salvas.\n"
        "Use /relatorio para gerar relatório em Excel.\n"
        "Use /editar para editar uma aposta."
    )

# --- COMANDO /nova ---
async def nova(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Digite a data da aposta (DD/MM/AAAA):")
    user_data[update.effective_chat.id] = {}
    return VALOR

async def get_data_aposta(update: Update, context: ContextTypes.DEFAULT_TYPE):
    texto = update.message.text
    try:
        datetime.strptime(texto, "%d/%m/%Y")
        user_data[update.effective_chat.id]["data_aposta"] = texto
        await update.message.reply_text("Digite a data prevista para o resultado (DD/MM/AAAA):")
        return DATA_RESULTADO
    except:
        await update.message.reply_text("Data inválida. Use o formato DD/MM/AAAA:")
        return VALOR

async def get_data_resultado(update: Update, context: ContextTypes.DEFAULT_TYPE):
    texto = update.message.text
    try:
        datetime.strptime(texto, "%d/%m/%Y")
        user_data[update.effective_chat.id]["data_resultado"] = texto
        await update.message.reply_text("Quantas casas de apostas? (2 ou 3)")
        return NUM_CASAS
    except:
        await update.message.reply_text("Data inválida. Use o formato DD/MM/AAAA:")
        return DATA_RESULTADO

async def get_num_casas(update: Update, context: ContextTypes.DEFAULT_TYPE):
    texto = update.message.text
    if texto not in ["2", "3"]:
        await update.message.reply_text("Digite 2 ou 3:")
        return NUM_CASAS
    user_data[update.effective_chat.id]["num_casas"] = int(texto)
    await update.message.reply_text(f"Digite as casas de apostas separadas por vírgula:")
    return CASAS

async def get_casas(update: Update, context: ContextTypes.DEFAULT_TYPE):
    texto = update.message.text
    casas = [c.strip() for c in texto.split(",")]
    num_casas = user_data[update.effective_chat.id]["num_casas"]
    if len(casas) != num_casas:
        await update.message.reply_text(f"Envie exatamente {num_casas} casas, separadas por vírgula:")
        return CASAS
    user_data[update.effective_chat.id]["casas"] = casas
    await update.message.reply_text(f"Digite as odds separadas por vírgula:")
    return ODDS

async def get_odds(update: Update, context: ContextTypes.DEFAULT_TYPE):
    texto = update.message.text
    odds_text = [o.strip().replace(",", ".") for o in texto.split(",")]
    try:
        odds = [float(o) for o in odds_text]
    except:
        await update.message.reply_text("Odds inválidas. Digite números separados por vírgula:")
        return ODDS
    num_casas = user_data[update.effective_chat.id]["num_casas"]
    if len(odds) != num_casas:
        await update.message.reply_text(f"Envie exatamente {num_casas} odds, separadas por vírgula:")
        return ODDS
    user_data[update.effective_chat.id]["odds"] = odds
    await update.message.reply_text(f"Digite o valor investido (ex: 100):")
    return VALOR_INVESTIDO

async def get_valor_investido(update: Update, context: ContextTypes.DEFAULT_TYPE):
    texto = update.message.text.replace(",", ".")
    try:
        valor = float(texto)
        if valor <= 0:
            raise ValueError()
    except:
        await update.message.reply_text("Valor inválido. Digite um número maior que zero:")
        return VALOR_INVESTIDO

    data_aposta = user_data[update.effective_chat.id]["data_aposta"]
    data_resultado = user_data[update.effective_chat.id]["data_resultado"]
    casas = user_data[update.effective_chat.id]["casas"]
    odds = user_data[update.effective_chat.id]["odds"]
    num_casas = user_data[update.effective_chat.id]["num_casas"]

    valor_investido = valor
    lucros = calcular_lucro(odds, valor_investido)
    lucro_total = max(lucros)

    # Salva na planilha
    file_name = "surebets.xlsx"
    wb = openpyxl.load_workbook(file_name)
    ws = wb["SureBets"]

    linha = [
        data_aposta, data_resultado,
        casas[0] if num_casas >= 1 else None,
        odds[0] if num_casas >= 1 else None,
        casas[1] if num_casas >= 2 else None,
        odds[1] if num_casas >= 2 else None,
        casas[2] if num_casas == 3 else None,
        odds[2] if num_casas == 3 else None,
        lucros[0] if num_casas >= 1 else None,
        lucros[1] if num_casas >= 2 else None,
        lucros[2] if num_casas == 3 else None,
        valor_investido,
        lucro_total
    ]

    ws.append(linha)
    wb.save(file_name)

    # Monta mensagem para enviar
    aposta = linha
    texto, negativo = formatar_mensagem(aposta)

    await update.message.reply_text(texto, parse_mode="Markdown")

    return ConversationHandler.END

# --- COMANDO /historico ---
async def historico(update: Update, context: ContextTypes.DEFAULT_TYPE):
    file_name = "surebets.xlsx"
    if not os.path.exists(file_name):
        await update.message.reply_text("Nenhuma aposta cadastrada ainda.")
        return

    wb = openpyxl.load_workbook(file_name)
    ws = wb["SureBets"]

    mensagens = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        texto, negativo = formatar_mensagem(row)
        mensagens.append(texto)

    if not mensagens:
        await update.message.reply_text("Nenhuma aposta cadastrada ainda.")
        return

    for msg in mensagens:
        await update.message.reply_text(msg, parse_mode="Markdown")

# --- COMANDO /relatorio ---
async def relatorio(update: Update, context: ContextTypes.DEFAULT_TYPE):
    file_name = "surebets.xlsx"
    if not os.path.exists(file_name):
        await update.message.reply_text("Nenhuma aposta cadastrada para gerar relatório.")
        return

    wb = openpyxl.load_workbook(file_name)
    ws = wb["SureBets"]

    datas = []
    lucros = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        datas.append(row[0])
        lucros.append(row[12])

    # Gera gráfico
    plt.figure(figsize=(8,4))
    plt.bar(datas, lucros, color=["green" if l >= 0 else "red" for l in lucros])
    plt.xticks(rotation=45)
    plt.title("Lucro total por aposta")
    plt.tight_layout()

    img_bytes = BytesIO()
    plt.savefig(img_bytes, format="png")
    img_bytes.seek(0)

    await update.message.reply_photo(photo=img_bytes)

    # Envia o arquivo Excel
    await update.message.reply_document(document=InputFile(file_name), filename=file_name)

# --- COMANDO /editar ---
async def editar(update: Update, context: ContextTypes.DEFAULT_TYPE):
    file_name = "surebets.xlsx"
    if not os.path.exists(file_name):
        await update.message.reply_text("Nenhuma aposta cadastrada para editar.")
        return ConversationHandler.END

    wb = openpyxl.load_workbook(file_name)
    ws = wb["SureBets"]
    apostas = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        resumo = f"{idx-1}: {row[0]} - Casas: {', '.join([c for c in [row[2], row[4], row[6]] if c])}"
        apostas.append((idx, resumo))

    teclado = [
        [InlineKeyboardButton(resumo, callback_data=str(idx))] for idx, resumo in apostas
    ]

    await update.message.reply_text("Escolha a aposta para editar:", reply_markup=InlineKeyboardMarkup(teclado))
    user_data[update.effective_chat.id] = {"apostas": apostas}
    return EDIT_SELECTION

async def edit_selection(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    escolha = int(query.data)
    user_data[query.message.chat_id]["edit_index"] = escolha - 2  # índice 0-based na planilha (linha -2)
    await query.message.reply_text(
        "Qual campo deseja editar?\n"
        "Opções: data_aposta, data_resultado, casa1, odd1, casa2, odd2, casa3, odd3, valor_investido"
    )
    return EDIT_FIELD

async def edit_field(update: Update, context: ContextTypes.DEFAULT_TYPE):
    campo = update.message.text.strip().lower()
    campos_validos = {
        "data_aposta": 0, "data_resultado": 1,
        "casa1": 2, "odd1": 3,
        "casa2": 4, "odd2": 5,
        "casa3": 6, "odd3": 7,
        "valor_investido": 11
    }
    if campo not in campos_validos:
        await update.message.reply_text("Campo inválido, tente novamente:")
        return EDIT_FIELD
    user_data[update.effective_chat.id]["edit_field"] = campo
    await update.message.reply_text(f"Digite o novo valor para {campo}:")
    return EDIT_VALUE

async def edit_value(update: Update, context: ContextTypes.DEFAULT_TYPE):
    valor = update.message.text.strip()
    idx = user_data[update.effective_chat.id]["edit_index"]
    campo = user_data[update.effective_chat.id]["edit_field"]
    campos_validos = {
        "data_aposta": 0, "data_resultado": 1,
        "casa1": 2, "odd1": 3,
        "casa2": 4, "odd2": 5,
        "casa3": 6, "odd3": 7,
        "valor_investido": 11
    }
    coluna = campos_validos[campo]

    wb = openpyxl.load_workbook("surebets.xlsx")
    ws = wb["SureBets"]
    linha = idx + 2  # corrige linha no Excel

    # Atualiza o valor e recalcula lucro se necessário
    if coluna in [3,5,7,11]:
        # campos numéricos
        try:
            if coluna == 11:
                valor_n = float(valor.replace(",", "."))
                ws.cell(row=linha, column=coluna+1, value=valor_n)
            else:
                valor_n = float(valor.replace(",", "."))
                ws.cell(row=linha, column=coluna+1, value=valor_n)
        except:
            await update.message.reply_text("Valor inválido para campo numérico, tente novamente:")
            return EDIT_VALUE
    else:
        ws.cell(row=linha, column=coluna+1, value=valor)

    # Recalcular lucros e lucro total
    odds = [
        ws.cell(row=linha, column=4).value,
        ws.cell(row=linha, column=6).value,
        ws.cell(row=linha, column=8).value,
    ]
    valor_investido = ws.cell(row=linha, column=12).value
    lucros = calcular_lucro(odds, valor_investido)

    ws.cell(row=linha, column=9, value=lucros[0])
    ws.cell(row=linha, column=10, value=lucros[1])
    ws.cell(row=linha, column=11, value=lucros[2])
    ws.cell(row=linha, column=13, value=max(lucros))

    wb.save("surebets.xlsx")

    await update.message.reply_text("Aposta atualizada com sucesso!")
    return ConversationHandler.END

# Notificação diária para apostas que batem no dia
async def notificacoes_diarias(app):
    file_name = "surebets.xlsx"
    if not os.path.exists(file_name):
        return

    wb = openpyxl.load_workbook(file_name)
    ws = wb["SureBets"]

    hoje = datetime.now().strftime("%d/%m/%Y")
    mensagens = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        data_resultado = row[1]
        lucro_total = row[12]
        if data_resultado == hoje:
            msg = f"Aposta com resultado hoje ({hoje}):\n"
            msg += f"Lucro esperado: R$ {lucro_total:.2f}\n"
            casas = [row[2], row[4], row[6]]
            odds = [row[3], row[5], row[7]]
            for i, casa in enumerate(casas):
                if casa:
                    msg += f"Casa {i+1}: {casa} (Odd: {odds[i]})\n"
            mensagens.append(msg)

    for msg in mensagens:
        await app.bot.send_message(chat_id=CHAT_ID, text=msg)

async def agendar_notificacoes(application):
    import asyncio
    while True:
        await notificacoes_diarias(application)
        await asyncio.sleep(86400)  # 24h

def main():
    criar_planilha()
    app = ApplicationBuilder().token(TOKEN).build()

    conv_handler = ConversationHandler(
        entry_points=[CommandHandler('nova', nova)],
        states={
            VALOR: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_data_aposta)],
            DATA_RESULTADO: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_data_resultado)],
            NUM_CASAS: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_num_casas)],
            CASAS: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_casas)],
            ODDS: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_odds)],
            VALOR_INVESTIDO: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_valor_investido)],
            EDIT_SELECTION: [CallbackQueryHandler(edit_selection)],
            EDIT_FIELD: [MessageHandler(filters.TEXT & ~filters.COMMAND, edit_field)],
            EDIT_VALUE: [MessageHandler(filters.TEXT & ~filters.COMMAND, edit_value)],
        },
        fallbacks=[]
    )

    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("historico", historico))
    app.add_handler(CommandHandler("relatorio", relatorio))
    app.add_handler(CommandHandler("editar", editar))
    app.add_handler(conv_handler)

    import asyncio
    app.job_queue.run_repeating(lambda ctx: asyncio.create_task(notificacoes_diarias(app)), interval=86400, first=10)

    app.run_polling()

if __name__ == "__main__":
    main()
