# Bot de Apostas - SureBets ESTILIZADO v2.0
# Versão com visual profissional e embeds do Discord

import os
import openpyxl
from datetime import datetime, time
from telegram import Update, InlineKeyboardMarkup, InlineKeyboardButton, InputFile
from telegram.ext import (
    ApplicationBuilder, CommandHandler, MessageHandler, filters,
    ContextTypes, ConversationHandler
)
import asyncio
import requests
import json

# ========== CONFIGURAÇÕES - ALTERE AQUI ==========
TOKEN = "7910538668:AAGJ2pcZ37WqKEYjsBSJwu_VzN8yzlI8iA4"
CHAT_ID = -1002881563925
DISCORD_WEBHOOK_URL = "https://discordapp.com/api/webhooks/1390157114525352046/9sbXdEfeU2SQv4lie3MLWpGqBrDF3gjnfRNdvW4M4qcd0cWcMr5qzC7m-xaruI_W3aRV" 
# ================================================

# Estados da conversa
(
    DATA_RESULTADO, HORARIO_RESULTADO, NUM_CASAS, CASAS, ODDS, VALOR_INVESTIDO,
    EDIT_SELECTION, EDIT_FIELD, EDIT_VALUE, APAGAR_SELECTION
) = range(10)

user_data = {}

def criar_planilha():
    """Cria a planilha Excel para armazenar as apostas"""
    file_name = "surebets.xlsx"
    if not os.path.exists(file_name):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "SureBets"
        ws.append([
            "Data Aposta", "Data Resultado", "Horario Resultado",
            "Casa 1", "Odd 1", "Valor Casa 1",
            "Casa 2", "Odd 2", "Valor Casa 2", 
            "Casa 3", "Odd 3", "Valor Casa 3",
            "Valor Total Investido", "Lucro Garantido", "ROI %"
        ])
        wb.save(file_name)
        print("📊 Planilha criada com sucesso!")

def calcular_surebet(odds, valor_total):
    """Calcula a distribuição de valores para surebet real"""
    soma_inversos = sum(1/odd for odd in odds if odd > 0)
    
    valores_casas = []
    lucros_por_casa = []
    
    for odd in odds:
        if odd > 0:
            valor_casa = (valor_total / odd) / soma_inversos
            valores_casas.append(round(valor_casa, 2))
            lucro_se_ganhar = (valor_casa * odd) - valor_total
            lucros_por_casa.append(round(lucro_se_ganhar, 2))
        else:
            valores_casas.append(0)
            lucros_por_casa.append(0)
    
    lucro_garantido = lucros_por_casa[0] if lucros_por_casa else 0
    roi_percentual = (lucro_garantido / valor_total * 100) if valor_total > 0 else 0
    
    return valores_casas, lucro_garantido, roi_percentual

def formatar_mensagem_surebet_telegram(aposta):
    """Formata mensagem elegante para o Telegram"""
    data_aposta = aposta[0]
    data_resultado = aposta[1]
    horario_resultado = aposta[2]
    
    casas = [aposta[3], aposta[6], aposta[9]]
    odds = [aposta[4], aposta[7], aposta[10]]
    valores_casas = [aposta[5], aposta[8], aposta[11]]
    
    valor_total = aposta[12]
    lucro_garantido = aposta[13]
    roi_percentual = aposta[14]

    # Cabeçalho elegante
    texto = "🎯 *SUREBET FINALIZADA* 🎯\n"
    texto += "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
    
    # Informações principais
    texto += f"📅 *Data da Aposta:* `{data_aposta}`\n"
    texto += f"⏰ *Resultado:* `{data_resultado}` às `{horario_resultado}`\n\n"
    
    # Resumo financeiro
    texto += "💰 *RESUMO FINANCEIRO*\n"
    texto += f"├ 💵 Valor Investido: `R$ {valor_total:.2f}`\n"
    texto += f"├ 🎯 Lucro Garantido: `R$ {lucro_garantido:.2f}`\n"
    texto += f"└ 📈 ROI: `{roi_percentual:.2f}%`\n\n"
    
    # Detalhes das casas
    texto += "🏠 *DISTRIBUIÇÃO DAS APOSTAS*\n"
    for i in range(3):
        if casas[i] and valores_casas[i] and valores_casas[i] > 0:
            texto += f"├ *{casas[i]}*\n"
            texto += f"│  ├ Odd: `{odds[i]}`\n"
            texto += f"│  └ Valor: `R$ {valores_casas[i]:.2f}`\n"
    
    texto += "\n━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
    texto += "✅ *Lucro garantido independente do resultado!*"
    
    return texto

def criar_embed_discord(aposta):
    """Cria embed elegante para o Discord"""
    data_aposta = aposta[0]
    data_resultado = aposta[1]
    horario_resultado = aposta[2]
    
    casas = [aposta[3], aposta[6], aposta[9]]
    odds = [aposta[4], aposta[7], aposta[10]]
    valores_casas = [aposta[5], aposta[8], aposta[11]]
    
    valor_total = aposta[12]
    lucro_garantido = aposta[13]
    roi_percentual = aposta[14]
    
    # Determinar cor baseada no ROI
    if roi_percentual >= 30:
        color = 0x00FF00  # Verde brilhante
    elif roi_percentual >= 15:
        color = 0xFFFF00  # Amarelo
    else:
        color = 0xFF6600  # Laranja
    
    embed = {
        "title": "🎯 SUREBET FINALIZADA",
        "description": f"**Aposta realizada em {data_aposta}**",
        "color": color,
        "timestamp": datetime.now().isoformat(),
        "fields": [
            {
                "name": "📅 Data do Resultado",
                "value": f"`{data_resultado}` às `{horario_resultado}`",
                "inline": True
            },
            {
                "name": "💰 Valor Investido",
                "value": f"```R$ {valor_total:.2f}```",
                "inline": True
            },
            {
                "name": "🎯 Lucro Garantido",
                "value": f"```R$ {lucro_garantido:.2f}```",
                "inline": True
            },
            {
                "name": "📈 ROI",
                "value": f"```{roi_percentual:.2f}%```",
                "inline": True
            }
        ],
        "footer": {
            "text": "SureBets Bot • Lucro garantido independente do resultado",
            "icon_url": "https://cdn-icons-png.flaticon.com/512/3135/3135715.png"
        }
    }
    
    # Adicionar campos das casas
    casas_info = ""
    for i in range(3):
        if casas[i] and valores_casas[i] and valores_casas[i] > 0:
            casas_info += f"**{casas[i]}**\n"
            casas_info += f"├ Odd: `{odds[i]}`\n"
            casas_info += f"└ Apostar: `R$ {valores_casas[i]:.2f}`\n\n"
    
    if casas_info:
        embed["fields"].append({
            "name": "🏠 Distribuição das Apostas",
            "value": casas_info,
            "inline": False
        })
    
    return embed

async def send_discord_embed(embed_data):
    """Envia embed elegante para o Discord"""
    if DISCORD_WEBHOOK_URL:
        payload = {
            "embeds": [embed_data]
        }
        try:
            response = requests.post(DISCORD_WEBHOOK_URL, json=payload)
            response.raise_for_status()
            print("✅ Embed enviado para o Discord!")
        except requests.exceptions.RequestException as e:
            print(f"❌ Erro ao enviar embed para Discord: {e}")
    else:
        print("⚠️ DISCORD_WEBHOOK_URL não configurado.")

# ========== COMANDOS DO BOT ==========

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Comando /start - Apresentação elegante do bot"""
    mensagem = """🎯 *BOT DE SUREBETS ESTILIZADO v2.0* 🎯
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

🤖 *Bem-vindo ao sistema mais avançado de SureBets!*

Eu ajudo você a gerenciar suas apostas com:
✅ Cálculos precisos de distribuição
✅ Notificações automáticas por horário
✅ Relatórios detalhados em Excel
✅ Interface elegante e profissional

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

📋 *COMANDOS DISPONÍVEIS:*

🆕 `/nova` - Criar nova surebet
📊 `/historico` - Ver histórico completo
📈 `/relatorio` - Gerar relatório Excel
🗑️ `/apagar` - Remover aposta do histórico
🔔 `/testar_notif` - Testar notificações
❌ `/cancelar` - Cancelar operação atual

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

🚀 *NOVIDADES v2.0:*
✨ Interface completamente redesenhada
🎨 Embeds elegantes no Discord
📱 Mensagens mais organizadas
🔔 Sistema de notificações aprimorado

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

💡 *Digite* `/nova` *para começar sua primeira surebet!*"""
    
    await update.message.reply_text(mensagem, parse_mode='Markdown')

async def nova(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Comando /nova - Início elegante do processo"""
    mensagem = """🆕 *CRIANDO NOVA SUREBET* 🆕
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

🎯 Vamos calcular sua próxima surebet com precisão!

📅 *Primeiro, me informe:*
**Qual a data do resultado da aposta?**

📝 *Formato:* `DD/MM/AAAA`
💡 *Exemplo:* `15/07/2025`

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━"""
    
    await update.message.reply_text(mensagem, parse_mode='Markdown')
    return DATA_RESULTADO

async def get_data_resultado(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Coleta a data do resultado com estilo"""
    user_data[update.effective_user.id] = {"data_resultado": update.message.text}
    
    mensagem = """⏰ *HORÁRIO DO RESULTADO* ⏰
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

🕐 **Qual o horário exato do resultado?**

📝 *Formato:* `HH:MM`
💡 *Exemplo:* `18:30`

🔔 *Importante:* Este será o horário da notificação automática!

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━"""
    
    await update.message.reply_text(mensagem, parse_mode='Markdown')
    return HORARIO_RESULTADO

async def get_horario_resultado(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Coleta o horário com estilo"""
    user_data[update.effective_user.id]["horario_resultado"] = update.message.text
    
    mensagem = """🏠 *CASAS DE APOSTAS* 🏠
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

🎯 **Quantas casas de apostas você vai usar?**

📊 *Opções disponíveis:*
├ `2` casas - Surebet simples
└ `3` casas - Surebet tripla

💡 *Recomendado:* 2-3 casas para máxima eficiência

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━"""
    
    await update.message.reply_text(mensagem, parse_mode='Markdown')
    return NUM_CASAS

async def get_num_casas(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Coleta número de casas com validação elegante"""
    try:
        num_casas = int(update.message.text)
        if not 2 <= num_casas <= 3:
            mensagem = """❌ *NÚMERO INVÁLIDO* ❌
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

⚠️ Para surebets eficientes, você precisa de **2 ou 3 casas**.

📊 *Opções válidas:*
├ `2` - Surebet com duas casas
└ `3` - Surebet com três casas

🔄 *Tente novamente:*"""
            
            await update.message.reply_text(mensagem, parse_mode='Markdown')
            return NUM_CASAS
        
        user_data[update.effective_user.id]["num_casas"] = num_casas
        user_data[update.effective_user.id]["casas"] = []
        user_data[update.effective_user.id]["odds"] = []
        
        mensagem = f"""🏠 *NOMES DAS CASAS* 🏠
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

📝 **Nome da Casa 1:**

💡 *Exemplos:* Bet365, Betano, Sportingbet, etc.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━"""
        
        await update.message.reply_text(mensagem, parse_mode='Markdown')
        return CASAS
        
    except ValueError:
        mensagem = """❌ *ENTRADA INVÁLIDA* ❌
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

⚠️ Por favor, digite apenas um **número inteiro**.

📊 *Opções válidas:*
├ `2` - Duas casas de apostas
└ `3` - Três casas de apostas

🔄 *Tente novamente:*"""
        
        await update.message.reply_text(mensagem, parse_mode='Markdown')
        return NUM_CASAS

async def get_casas(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Coleta nomes das casas com progresso visual"""
    user_id = update.effective_user.id
    user_data[user_id]["casas"].append(update.message.text)
    
    num_casas = user_data[user_id]["num_casas"]
    casas_coletadas = len(user_data[user_id]["casas"])
    
    if casas_coletadas < num_casas:
        # Mostrar progresso
        progresso = "█" * casas_coletadas + "░" * (num_casas - casas_coletadas)
        
        mensagem = f"""🏠 *NOMES DAS CASAS* 🏠
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

📊 *Progresso:* `{progresso}` ({casas_coletadas}/{num_casas})

✅ *Casas já adicionadas:*
{chr(10).join([f"├ {casa}" for casa in user_data[user_id]["casas"]])}

📝 **Nome da Casa {casas_coletadas + 1}:**

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━"""
        
        await update.message.reply_text(mensagem, parse_mode='Markdown')
        return CASAS
    else:
        # Todas as casas coletadas, partir para odds
        casas_lista = "\n".join([f"├ {casa}" for casa in user_data[user_id]["casas"]])
        
        mensagem = f"""📊 *ODDS DAS CASAS* 📊
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

✅ *Casas configuradas:*
{casas_lista}

🎯 **Odd da Casa 1:** `{user_data[user_id]["casas"][0]}`

📝 *Formato:* Número decimal (ex: `2.50`)
💡 *Importante:* Use ponto para decimais

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━"""
        
        await update.message.reply_text(mensagem, parse_mode='Markdown')
        return ODDS

async def get_odds(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Coleta odds com validação e progresso visual"""
    user_id = update.effective_user.id
    try:
        odd = float(update.message.text)
        if odd <= 1:
            mensagem = """❌ *ODD INVÁLIDA* ❌
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

⚠️ A odd deve ser **maior que 1.00**

📊 *Exemplos válidos:*
├ `1.50` - Odd baixa
├ `2.75` - Odd média  
└ `5.20` - Odd alta

🔄 *Tente novamente:*"""
            
            await update.message.reply_text(mensagem, parse_mode='Markdown')
            return ODDS
            
        user_data[user_id]["odds"].append(odd)
        
        num_casas = user_data[user_id]["num_casas"]
        odds_coletadas = len(user_data[user_id]["odds"])
        
        if odds_coletadas < num_casas:
            # Mostrar progresso das odds
            progresso = "█" * odds_coletadas + "░" * (num_casas - odds_coletadas)
            
            # Mostrar odds já coletadas
            odds_lista = ""
            for i, (casa, odd_valor) in enumerate(zip(user_data[user_id]["casas"], user_data[user_id]["odds"])):
                odds_lista += f"├ {casa}: `{odd_valor}`\n"
            
            mensagem = f"""📊 *ODDS DAS CASAS* 📊
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

📊 *Progresso:* `{progresso}` ({odds_coletadas}/{num_casas})

✅ *Odds já coletadas:*
{odds_lista}

🎯 **Odd da Casa {odds_coletadas + 1}:** `{user_data[user_id]["casas"][odds_coletadas]}`

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━"""
            
            await update.message.reply_text(mensagem, parse_mode='Markdown')
            return ODDS
        else:
            # Todas as odds coletadas, verificar surebet
            odds = user_data[user_id]["odds"]
            soma_inversos = sum(1/odd for odd in odds)
            
            # Mostrar análise da surebet
            if soma_inversos >= 1:
                status = "❌ NÃO É SUREBET"
                cor = "🔴"
                aviso = f"""
⚠️ **ATENÇÃO:** Essas odds não formam uma surebet válida!

📊 *Análise técnica:*
├ Soma dos inversos: `{soma_inversos:.4f}`
├ Para ser surebet: `< 1.0000`
└ Status: {status}

🤔 *Deseja continuar mesmo assim?*"""
            else:
                status = "✅ SUREBET VÁLIDA"
                cor = "🟢"
                roi_estimado = ((1/soma_inversos - 1) * 100)
                aviso = f"""
🎉 **PARABÉNS!** Você tem uma surebet válida!

📊 *Análise técnica:*
├ Soma dos inversos: `{soma_inversos:.4f}`
├ ROI estimado: `~{roi_estimado:.2f}%`
└ Status: {status}"""
            
            mensagem = f"""🔍 *ANÁLISE DA SUREBET* 🔍
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

{cor} **RESULTADO DA ANÁLISE**
{aviso}

💰 **Agora, qual o valor total que você quer investir?**

📝 *Formato:* Número decimal (ex: `100.00`)
💡 *Este valor será distribuído automaticamente*

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━"""
            
            await update.message.reply_text(mensagem, parse_mode='Markdown')
            return VALOR_INVESTIDO
            
    except ValueError:
        mensagem = """❌ *FORMATO INVÁLIDO* ❌
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

⚠️ Por favor, insira um **número válido** para a odd.

📊 *Formatos aceitos:*
├ `2.5` ou `2.50`
├ `1.75`
└ `3.25`

🚫 *Não use:* vírgulas, letras ou símbolos

🔄 *Tente novamente:*"""
        
        await update.message.reply_text(mensagem, parse_mode='Markdown')
        return ODDS

async def get_valor_investido(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Finaliza a surebet com cálculos e apresentação elegante"""
    user_id = update.effective_user.id
    try:
        valor_total = float(update.message.text)
        if valor_total <= 0:
            mensagem = """❌ *VALOR INVÁLIDO* ❌
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

⚠️ O valor deve ser **maior que zero**.

💰 *Exemplos válidos:*
├ `50.00` - Cinquenta reais
├ `100.50` - Cem reais e cinquenta centavos
└ `1000` - Mil reais

🔄 *Tente novamente:*"""
            
            await update.message.reply_text(mensagem, parse_mode='Markdown')
            return VALOR_INVESTIDO
            
        user_data[user_id]["valor_total"] = valor_total
        
        # Calcular a distribuição da surebet
        odds = user_data[user_id]["odds"]
        valores_casas, lucro_garantido, roi_percentual = calcular_surebet(odds, valor_total)
        
        # Salvar no Excel
        file_name = "surebets.xlsx"
        wb = openpyxl.load_workbook(file_name)
        ws = wb["SureBets"]

        data_aposta = datetime.now().strftime("%d/%m/%Y")
        data_resultado = user_data[user_id]["data_resultado"]
        horario_resultado = user_data[user_id]["horario_resultado"]
        casas = user_data[user_id]["casas"]

        # Preparar dados para a planilha
        row_data = [data_aposta, data_resultado, horario_resultado]
        
        # Adicionar dados das casas (máximo 3)
        for i in range(3):
            if i < len(casas):
                row_data.extend([casas[i], odds[i], valores_casas[i]])
            else:
                row_data.extend(["", "", ""])
        
        # Adicionar totais
        row_data.extend([valor_total, lucro_garantido, roi_percentual])

        ws.append(row_data)
        wb.save(file_name)

        # Determinar emoji baseado no ROI
        if roi_percentual >= 30:
            emoji_roi = "🚀"
        elif roi_percentual >= 15:
            emoji_roi = "📈"
        elif roi_percentual >= 5:
            emoji_roi = "💹"
        else:
            emoji_roi = "📊"

        # Mensagem de sucesso elegante
        mensagem_sucesso = f"""🎉 *SUREBET CRIADA COM SUCESSO!* 🎉
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

✅ **RESUMO DA OPERAÇÃO**

📅 *Data da Aposta:* `{data_aposta}`
⏰ *Resultado:* `{data_resultado}` às `{horario_resultado}`

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

💰 **ANÁLISE FINANCEIRA**
├ 💵 Valor Total: `R$ {valor_total:.2f}`
├ 🎯 Lucro Garantido: `R$ {lucro_garantido:.2f}`
└ {emoji_roi} ROI: `{roi_percentual:.2f}%`

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

🏠 **COMO APOSTAR:**
"""
        
        for i, casa in enumerate(casas):
            mensagem_sucesso += f"""
🎯 **{casa}**
├ Odd: `{odds[i]}`
├ Apostar: `R$ {valores_casas[i]:.2f}`
└ Retorno: `R$ {valores_casas[i] * odds[i]:.2f}`"""
        
        mensagem_sucesso += f"""

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

✨ **INDEPENDENTE DO RESULTADO:**
🎯 Você terá `R$ {lucro_garantido:.2f}` de lucro garantido!

🔔 **NOTIFICAÇÃO AGENDADA:**
📅 `{data_resultado}` às `{horario_resultado}`

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

🚀 *Surebet salva com sucesso no sistema!*"""
        
        await update.message.reply_text(mensagem_sucesso, parse_mode='Markdown')
        
        # Limpar dados do usuário
        if user_id in user_data:
            del user_data[user_id]
        
        return ConversationHandler.END
        
    except ValueError:
        mensagem = """❌ *FORMATO INVÁLIDO* ❌
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

⚠️ Por favor, insira um **número válido** para o valor.

💰 *Formatos aceitos:*
├ `100` - Cem reais
├ `50.75` - Cinquenta e setenta e cinco centavos
└ `1000.00` - Mil reais

🚫 *Não use:* vírgulas, símbolos ou letras

🔄 *Tente novamente:*"""
        
        await update.message.reply_text(mensagem, parse_mode='Markdown')
        return VALOR_INVESTIDO

async def historico(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Comando /historico - Exibe histórico elegante"""
    file_name = "surebets.xlsx"
    if not os.path.exists(file_name):
        mensagem = """📋 *HISTÓRICO VAZIO* 📋
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

📊 Nenhuma surebet encontrada no sistema.

💡 *Para começar:*
├ Digite `/nova` para criar sua primeira surebet
└ O histórico aparecerá aqui automaticamente

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━"""
        
        await update.message.reply_text(mensagem, parse_mode='Markdown')
        return

    wb = openpyxl.load_workbook(file_name)
    ws = wb["SureBets"]

    if ws.max_row <= 1:
        mensagem = """📋 *HISTÓRICO VAZIO* 📋
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

📊 Nenhuma surebet encontrada no sistema.

💡 *Para começar:*
├ Digite `/nova` para criar sua primeira surebet
└ O histórico aparecerá aqui automaticamente

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━"""
        
        await update.message.reply_text(mensagem, parse_mode='Markdown')
        return

    # Cabeçalho do histórico
    message = """📋 *HISTÓRICO DE SUREBETS* 📋
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

"""
    
    total_investido = 0
    total_lucro = 0
    
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
        data_aposta = row[0]
        data_resultado = row[1]
        horario_resultado = row[2]
        valor_total = row[12] if row[12] else 0
        lucro_garantido = row[13] if row[13] else 0
        roi_percentual = row[14] if row[14] else 0
        
        total_investido += valor_total
        total_lucro += lucro_garantido
        
        # Emoji baseado no ROI
        if roi_percentual >= 30:
            status_emoji = "🚀"
        elif roi_percentual >= 15:
            status_emoji = "📈"
        elif roi_percentual >= 5:
            status_emoji = "💹"
        else:
            status_emoji = "📊"
        
        message += f"""{status_emoji} **Surebet #{row_num}**
├ 📅 `{data_aposta}` → `{data_resultado} {horario_resultado}`
├ 💰 Investido: `R$ {valor_total:.2f}`
├ 🎯 Lucro: `R$ {lucro_garantido:.2f}`
└ 📈 ROI: `{roi_percentual:.1f}%`

"""
    
    # Resumo final elegante
    roi_medio = (total_lucro/total_investido*100) if total_investido > 0 else 0
    
    # Emoji para o resumo geral
    if roi_medio >= 25:
        resumo_emoji = "🏆"
    elif roi_medio >= 15:
        resumo_emoji = "🥇"
    elif roi_medio >= 10:
        resumo_emoji = "🥈"
    else:
        resumo_emoji = "🥉"
    
    message += f"""━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

{resumo_emoji} **RESUMO GERAL**
├ 💰 Total Investido: `R$ {total_investido:.2f}`
├ 🎯 Lucro Total: `R$ {total_lucro:.2f}`
├ 📈 ROI Médio: `{roi_medio:.1f}%`
└ 📊 Total de Surebets: `{ws.max_row - 1}`

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━"""
    
    await update.message.reply_text(message, parse_mode='Markdown')

async def apagar(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Comando /apagar - Interface elegante para remoção"""
    file_name = "surebets.xlsx"
    if not os.path.exists(file_name):
        mensagem = """🗑️ *NADA PARA APAGAR* 🗑️
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

📊 Nenhuma surebet encontrada no sistema.

💡 *Para ter surebets para apagar:*
├ Digite `/nova` para criar surebets
└ Depois use `/apagar` para removê-las

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━"""
        
        await update.message.reply_text(mensagem, parse_mode='Markdown')
        return ConversationHandler.END

    wb = openpyxl.load_workbook(file_name)
    ws = wb["SureBets"]

    if ws.max_row <= 1:
        mensagem = """🗑️ *NADA PARA APAGAR* 🗑️
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

📊 Nenhuma surebet encontrada no sistema.

💡 *Para ter surebets para apagar:*
├ Digite `/nova` para criar surebets
└ Depois use `/apagar` para removê-las

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━"""
        
        await update.message.reply_text(mensagem, parse_mode='Markdown')
        return ConversationHandler.END

    message = """🗑️ *REMOVER SUREBET* 🗑️
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

⚠️ **Selecione a surebet que deseja remover:**

"""
    
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
        data_aposta = row[0]
        data_resultado = row[1]
        horario_resultado = row[2]
        valor_total = row[12] if row[12] else 0
        lucro_garantido = row[13] if row[13] else 0
        roi_percentual = row[14] if row[14] else 0
        
        # Emoji baseado no ROI
        if roi_percentual >= 30:
            status_emoji = "🚀"
        elif roi_percentual >= 15:
            status_emoji = "📈"
        else:
            status_emoji = "📊"
        
        message += f"""{status_emoji} **#{row_num}** `{data_aposta}` → `{data_resultado} {horario_resultado}`
├ 💰 `R$ {valor_total:.2f}` | 🎯 `R$ {lucro_garantido:.2f}`

"""
    
    message += """━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

🔢 **Digite o número da surebet para remover:**
💡 *Exemplo:* `1` para remover a primeira surebet

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━"""
    
    await update.message.reply_text(message, parse_mode='Markdown')
    return APAGAR_SELECTION

async def get_apagar_selection(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Processa remoção com confirmação elegante"""
    try:
        row_to_delete = int(update.message.text)
        
        file_name = "surebets.xlsx"
        wb = openpyxl.load_workbook(file_name)
        ws = wb["SureBets"]
        
        if row_to_delete < 1 or row_to_delete > (ws.max_row - 1):
            mensagem = f"""❌ *NÚMERO INVÁLIDO* ❌
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

⚠️ Número `{row_to_delete}` não existe na lista.

📊 *Números válidos:* `1` até `{ws.max_row - 1}`

🔄 *Tente novamente com um número válido:*

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━"""
            
            await update.message.reply_text(mensagem, parse_mode='Markdown')
            return APAGAR_SELECTION
        
        # Obter dados da linha antes de apagar
        row_data = list(ws.iter_rows(min_row=row_to_delete + 1, max_row=row_to_delete + 1, values_only=True))[0]
        data_aposta = row_data[0]
        data_resultado = row_data[1]
        horario_resultado = row_data[2]
        valor_total = row_data[12] if row_data[12] else 0
        lucro_garantido = row_data[13] if row_data[13] else 0
        roi_percentual = row_data[14] if row_data[14] else 0
        
        # Apagar a linha
        ws.delete_rows(row_to_delete + 1)
        wb.save(file_name)
        
        # Emoji baseado no ROI da surebet removida
        if roi_percentual >= 30:
            emoji_removida = "🚀"
        elif roi_percentual >= 15:
            emoji_removida = "📈"
        else:
            emoji_removida = "📊"
        
        mensagem_confirmacao = f"""✅ *SUREBET REMOVIDA COM SUCESSO!* ✅
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

🗑️ **SUREBET REMOVIDA:**

{emoji_removida} **Surebet #{row_to_delete}**
├ 📅 Data: `{data_aposta}` → `{data_resultado} {horario_resultado}`
├ 💰 Valor: `R$ {valor_total:.2f}`
├ 🎯 Lucro: `R$ {lucro_garantido:.2f}`
└ 📈 ROI: `{roi_percentual:.2f}%`

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

🔥 **A surebet foi removida permanentemente do sistema.**

💡 *Para ver o histórico atualizado, use* `/historico`

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━"""
        
        await update.message.reply_text(mensagem_confirmacao, parse_mode='Markdown')
        return ConversationHandler.END
        
    except ValueError:
        mensagem = """❌ *FORMATO INVÁLIDO* ❌
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

⚠️ Por favor, digite apenas o **número** da surebet.

📊 *Formato correto:*
├ `1` - Para remover a primeira surebet
├ `2` - Para remover a segunda surebet
└ `3` - Para remover a terceira surebet

🚫 *Não use:* letras, símbolos ou espaços

🔄 *Tente novamente:*

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━"""
        
        await update.message.reply_text(mensagem, parse_mode='Markdown')
        return APAGAR_SELECTION
    except Exception as e:
        mensagem = f"""❌ *ERRO INESPERADO* ❌
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

⚠️ Ocorreu um erro ao remover a surebet:
`{str(e)}`

🔄 *Tente novamente ou use* `/cancelar`

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━"""
        
        await update.message.reply_text(mensagem, parse_mode='Markdown')
        return ConversationHandler.END

async def relatorio(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Comando /relatorio - Gera relatório com apresentação elegante"""
    file_name = "surebets.xlsx"
    if not os.path.exists(file_name):
        mensagem = """📊 *RELATÓRIO INDISPONÍVEL* 📊
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

📋 Nenhuma surebet encontrada para gerar relatório.

💡 *Para gerar relatórios:*
├ Digite `/nova` para criar surebets
└ Use `/relatorio` para baixar o Excel

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━"""
        
        await update.message.reply_text(mensagem, parse_mode='Markdown')
        return

    try:
        # Mensagem de preparação
        mensagem_preparando = """📊 *GERANDO RELATÓRIO* 📊
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

⏳ **Preparando seu relatório completo...**

📋 *O arquivo Excel contém:*
├ ✅ Todas as surebets registradas
├ ✅ Cálculos detalhados de lucro
├ ✅ Análise de ROI por aposta
└ ✅ Dados organizados por data

🔄 *Aguarde um momento...*

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━"""
        
        await update.message.reply_text(mensagem_preparando, parse_mode='Markdown')
        
        # Enviar o arquivo
        with open(file_name, "rb") as excel_file:
            caption = """📊 *RELATÓRIO COMPLETO DE SUREBETS* 📊

✅ **Arquivo Excel gerado com sucesso!**

📋 *Este relatório contém:*
├ 📅 Histórico completo de apostas
├ 💰 Análise financeira detalhada  
├ 📈 Cálculos de ROI e lucros
└ 🏠 Dados de todas as casas utilizadas

💡 *Abra no Excel, Google Sheets ou LibreOffice*

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━"""
            
            await update.message.reply_document(
                document=InputFile(excel_file, filename=file_name), 
                caption=caption,
                parse_mode='Markdown'
            )
    except Exception as e:
        mensagem = f"""❌ *ERRO AO GERAR RELATÓRIO* ❌
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

⚠️ Ocorreu um erro ao gerar o relatório:
`{str(e)}`

🔄 *Tente novamente em alguns instantes.*

💡 *Se o problema persistir:*
├ Verifique se há surebets no histórico
└ Use `/historico` para verificar os dados

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━"""
        
        await update.message.reply_text(mensagem, parse_mode='Markdown')

async def testar_notif(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Comando /testar_notif - Interface elegante para testes"""
    try:
        mensagem_testando = """🔔 *TESTANDO NOTIFICAÇÕES* 🔔
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

⏳ **Verificando sistema de notificações...**

🔍 *Analisando:*
├ ⏰ Horário atual do sistema
├ 📅 Surebets agendadas para agora
├ 🔔 Status das notificações
└ 📱 Conectividade com Discord

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━"""
        
        await update.message.reply_text(mensagem_testando, parse_mode='Markdown')
        
        mensagens = await verificar_notificacoes()
        
        if mensagens:
            mensagem_encontradas = f"""✅ *NOTIFICAÇÕES ENCONTRADAS!* ✅
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

🎉 **Encontradas `{len(mensagens)}` notificações para enviar agora!**

🔔 *Enviando notificações...*

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━"""
            
            await update.message.reply_text(mensagem_encontradas, parse_mode='Markdown')
            
            for msg in mensagens:
                await update.message.reply_text(msg, parse_mode='Markdown')
        else:
            mensagem_vazio = """📋 *NENHUMA NOTIFICAÇÃO AGORA* 📋
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

⏰ **Nenhuma surebet para notificar no momento atual.**

✅ *Sistema funcionando corretamente!*

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━"""
            
            await update.message.reply_text(mensagem_vazio, parse_mode='Markdown')
            
        # Mostrar próximas notificações
        proximas = await verificar_proximas_notificacoes()
        if proximas:
            mensagem_proximas = f"""📅 *PRÓXIMAS NOTIFICAÇÕES* 📅
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

🔔 **Notificações agendadas:**

{proximas}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

💡 *As notificações serão enviadas automaticamente nos horários indicados.*

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━"""
            
            await update.message.reply_text(mensagem_proximas, parse_mode='Markdown')
        
    except Exception as e:
        mensagem_erro = f"""❌ *ERRO NO TESTE* ❌
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

⚠️ Ocorreu um erro ao testar as notificações:
`{str(e)}`

🔧 *Possíveis soluções:*
├ Verifique se há surebets cadastradas
├ Confirme se o sistema está funcionando
└ Tente novamente em alguns instantes

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━"""
        
        await update.message.reply_text(mensagem_erro, parse_mode='Markdown')

async def verificar_proximas_notificacoes():
    """Verifica próximas notificações com formatação elegante"""
    file_name = "surebets.xlsx"
    if not os.path.exists(file_name):
        return "📋 Nenhuma surebet cadastrada no sistema."

    wb = openpyxl.load_workbook(file_name)
    ws = wb["SureBets"]

    proximas = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        data_resultado = row[1]
        horario_resultado = row[2]
        casas = [row[3], row[6], row[9]]
        valor_total = row[12]
        roi_percentual = row[14] if row[14] else 0
        
        if data_resultado and horario_resultado:
            casas_str = ", ".join([casa for casa in casas if casa])
            
            # Emoji baseado no ROI
            if roi_percentual >= 30:
                emoji = "🚀"
            elif roi_percentual >= 15:
                emoji = "📈"
            else:
                emoji = "📊"
            
            proximas.append(f"{emoji} `{data_resultado}` às `{horario_resultado}` - {casas_str} (`R$ {valor_total:.2f}`)")
    
    return "\n".join(proximas) if proximas else "📋 Nenhuma notificação agendada."

async def cancelar(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Comando /cancelar - Cancelamento elegante"""
    user_id = update.effective_user.id
    if user_id in user_data:
        del user_data[user_id]
    
    mensagem = """❌ *OPERAÇÃO CANCELADA* ❌
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

🔄 **Operação cancelada com sucesso.**

💡 *Você pode:*
├ `/nova` - Criar nova surebet
├ `/historico` - Ver histórico
├ `/relatorio` - Gerar Excel
└ `/start` - Ver menu principal

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━"""
    
    await update.message.reply_text(mensagem, parse_mode='Markdown')
    return ConversationHandler.END

async def verificar_notificacoes():
    """Verifica notificações com logs elegantes"""
    file_name = "surebets.xlsx"
    if not os.path.exists(file_name):
        return []

    wb = openpyxl.load_workbook(file_name)
    ws = wb["SureBets"]

    now = datetime.now()
    today = now.strftime("%d/%m/%Y")
    current_time = now.strftime("%H:%M")
    
    print(f"🔍 Verificando notificações para {today} às {current_time}")
    
    mensagens = []

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
        data_resultado = row[1]
        horario_resultado = row[2]
        
        print(f"📋 Linha {row_num}: {data_resultado} às {horario_resultado}")
        
        if data_resultado == today and horario_resultado == current_time:
            print(f"✅ MATCH! Enviando notificação para linha {row_num}")
            texto_telegram = formatar_mensagem_surebet_telegram(row)
            mensagens.append(texto_telegram)
        else:
            print(f"❌ Não match: {data_resultado} != {today} ou {horario_resultado} != {current_time}")

    print(f"📤 Total de mensagens para enviar: {len(mensagens)}")
    return mensagens

async def notificacoes_por_horario(context: ContextTypes.DEFAULT_TYPE) -> None:
    """Executa notificações com embeds elegantes"""
    try:
        print("🔄 Executando verificação de notificações...")
        
        # Verificar notificações
        file_name = "surebets.xlsx"
        if not os.path.exists(file_name):
            return

        wb = openpyxl.load_workbook(file_name)
        ws = wb["SureBets"]

        now = datetime.now()
        today = now.strftime("%d/%m/%Y")
        current_time = now.strftime("%H:%M")
        
        print(f"🔍 Verificando para {today} às {current_time}")
        
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
            data_resultado = row[1]
            horario_resultado = row[2]
            
            if data_resultado == today and horario_resultado == current_time:
                print(f"✅ MATCH! Enviando notificação para linha {row_num}")
                
                # Enviar para Telegram
                if CHAT_ID:
                    try:
                        texto_telegram = formatar_mensagem_surebet_telegram(row)
                        await context.bot.send_message(chat_id=CHAT_ID, text=texto_telegram, parse_mode='Markdown')
                        print(f"✅ Notificação enviada para Telegram!")
                    except Exception as e:
                        print(f"❌ Erro ao enviar para Telegram: {e}")
                
                # Enviar embed para Discord
                try:
                    embed_discord = criar_embed_discord(row)
                    await send_discord_embed(embed_discord)
                    print(f"✅ Embed enviado para Discord!")
                except Exception as e:
                    print(f"❌ Erro ao enviar embed para Discord: {e}")
                
    except Exception as e:
        print(f"❌ Erro na verificação de notificações: {e}")

def main() -> None:
    """Função principal que inicia o bot"""
    try:
        print("Iniciando o Bot de SureBets FUNCIONAL...")
        
        if TOKEN == "SEU_TOKEN_DO_TELEGRAM_AQUI":
            print("ERRO: Token do Telegram nao configurado!")
            print("Por favor, substitua 'SEU_TOKEN_DO_TELEGRAM_AQUI' pelo seu token real.")
            print("Obtenha seu token em: @BotFather no Telegram")
            return
        
        criar_planilha()
        application = ApplicationBuilder().token(TOKEN).build()

        conv_handler_nova = ConversationHandler(
            entry_points=[CommandHandler("nova", nova)],
            states={
                DATA_RESULTADO: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_data_resultado)],
                HORARIO_RESULTADO: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_horario_resultado)],
                NUM_CASAS: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_num_casas)],
                CASAS: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_casas)],
                ODDS: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_odds)],
                VALOR_INVESTIDO: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_valor_investido)],
            },
            fallbacks=[CommandHandler("cancelar", cancelar)],
        )

        conv_handler_apagar = ConversationHandler(
            entry_points=[CommandHandler("apagar", apagar)],
            states={
                APAGAR_SELECTION: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_apagar_selection)],
            },
            fallbacks=[CommandHandler("cancelar", cancelar)],
        )

        application.add_handler(CommandHandler("start", start))
        application.add_handler(conv_handler_nova)
        application.add_handler(conv_handler_apagar)
        application.add_handler(CommandHandler("historico", historico))
        application.add_handler(CommandHandler("relatorio", relatorio))
        application.add_handler(CommandHandler("testar_notif", testar_notif))

        job_queue = application.job_queue
        job_queue.run_repeating(
            notificacoes_por_horario, 
            interval=60,
            first=10
        )

        print("Bot iniciado com sucesso!")
        print("Notificacoes configuradas para verificar a cada minuto")
        print("Use /testar_notif para testar o sistema")
        print("Pressione Ctrl+C para parar o bot")
        print("-" * 50)
        
        application.run_polling()
        
    except Exception as e:
        print(f"Erro ao iniciar o bot: {e}")
        print("Verifique se o token esta correto e se voce tem conexao com a internet.")

if __name__ == "__main__":
    main()

